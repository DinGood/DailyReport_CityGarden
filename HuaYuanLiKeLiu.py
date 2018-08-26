# -*- coding:utf-8 -*
import csv
import openpyxl
import datetime
import xlrd
import os

print('======================================')
print('>程序：花园里客流统计')
print('>作者：丁浒 / DinGood')
print('>邮件：dingood@qq.com')
print('>版本：v.18.8.26')
print('======================================')
print('请输入处理数据的日期（例：0824）：')

today = datetime.date.today()
yesterday = str(today - datetime.timedelta(days=1))
date = input()

def sumcsv(filename):
    sum_7_10 = 0
    sum_10_12 = 0
    sum_12_17 = 0
    sum_17_21 = 0
    sum_21_22 = 0

    with open(filename) as f:
        reader = csv.reader(f)
        for row in reader:
            if reader.line_num >= 9 and reader.line_num <= 11:
                sum_7_10 = sum_7_10 + int(row[1])
            if reader.line_num >= 12 and reader.line_num <= 13:
                sum_10_12 = sum_10_12 + int(row[1])
            if reader.line_num >= 14 and reader.line_num <= 18:
                sum_12_17 = sum_12_17 + int(row[1])
            if reader.line_num >= 19 and reader.line_num <= 22:
                sum_17_21 = sum_17_21 + int(row[1])
            if reader.line_num == 23:
                sum_21_22 = int(row[1])

    return sum_7_10, sum_10_12, sum_12_17, sum_17_21, sum_21_22


lnj710, lnj1012, lnj1217, lnj1721, lnj2122 = sumcsv("系统导出数据表/老娘舅入口_2018" + date + ".csv")
zrkz710, zrkz1012, zrkz1217, zrkz1721, zrkz2122 = sumcsv("系统导出数据表/主入口左_2018" + date + ".csv")
zrky710, zrky1012, zrky1217, zrky1721, zrky2122 = sumcsv("系统导出数据表/主入口右_2018" + date + ".csv")
xysjz710, xysjz1012, xysjz1217, xysjz1721, xysjz2122 = sumcsv("系统导出数据表/小杨生煎入口左_2018" + date + ".csv")
xysjy710, xysjy1012, xysjy1217, xysjy1721, xysjy2122 = sumcsv("系统导出数据表/小杨生煎入口右_2018" + date + ".csv")
dtft710, dtft1012, dtft1217, dtft1721, dtft2122 = sumcsv("系统导出数据表/地铁扶梯_2018" + date + ".csv")
dtrxt710, dtrxt1012, dtrxt1217, dtrxt1721, dtrxt2122 = sumcsv("系统导出数据表/地铁人行梯_2018" + date + ".csv")

# ------------------------------添加花园里读取环比数据--------------------------------------
workbook = xlrd.open_workbook('固定数据表/夫子庙2018年汇总 （1-12月）每日累加.xlsx')
sheets = workbook.sheet_names()
worksheet = workbook.sheet_by_name(sheets[15])  # 2018年汇总表中分表的页码
huanbi_date = str(today - datetime.timedelta(days=8))

if worksheet.cell_value(41, 9) > 0:
    huanbi = '上升'
else:
    huanbi = '下降'

# ------------------------------添加花园里读取环比数据--------------------------------------

wb = openpyxl.Workbook()
sheet = wb.active

sheet.merge_cells('A1:B1')
sheet.merge_cells('A2:B2')
sheet.merge_cells('A3:B3')
sheet.merge_cells('C1:D1')
sheet.merge_cells('C2:D2')
sheet.merge_cells('C3:D3')
sheet.merge_cells('E1:F1')
sheet.merge_cells('E2:F2')
sheet.merge_cells('E3:F3')
sheet.merge_cells('G1:H1')
sheet.merge_cells('G2:H2')
sheet.merge_cells('G3:H3')
sheet.merge_cells('I1:J1')
sheet.merge_cells('I2:J2')
sheet.merge_cells('I3:J3')


sheet.cell(1, 1, lnj710 + zrkz710 + zrky710)
sheet.cell(1, 3, lnj1012 + zrkz1012 + zrky1012)
sheet.cell(1, 5, lnj1217 + zrkz1217 + zrky1217)
sheet.cell(1, 7, lnj1721 + zrkz1721 + zrky1721)
sheet.cell(1, 9, lnj2122 + zrkz2122 + zrky2122)

sheet.cell(2, 1, xysjz710 + xysjy710)
sheet.cell(2, 3, xysjz1012 + xysjy1012)
sheet.cell(2, 5, xysjz1217 + xysjy1217)
sheet.cell(2, 7, xysjz1721 + xysjy1721)
sheet.cell(2, 9, xysjz2122 + xysjy2122)

sheet.cell(3, 1, dtft710 + dtrxt710)
sheet.cell(3, 3, dtft1012 + dtrxt1012)
sheet.cell(3, 5, dtft1217 + dtrxt1217)
sheet.cell(3, 7, dtft1721 + dtrxt1721)
sheet.cell(3, 9, dtft2122 + dtrxt2122)

sum_all = lnj710 + zrkz710 + zrky710 + lnj1012 + zrkz1012 + zrky1012 + lnj1217 + zrkz1217 +\
    zrky1217 + lnj1721 + zrkz1721 + zrky1721 + lnj2122 + zrkz2122 + zrky2122 + xysjz710 +\
    xysjy710 + xysjz1012 + xysjy1012 + xysjz1217 + xysjy1217 + xysjz1721 + xysjy1721 +\
    xysjz2122 + xysjy2122 + dtft710 + dtrxt710 + dtft1012 + dtrxt1012 + dtft1217 + dtrxt1217 +\
    dtft1721 + dtrxt1721 + dtft2122 + dtrxt2122

sum_710 = lnj710 + zrkz710 + zrky710 + xysjz710 + xysjy710 + dtft710 + dtrxt710
sum_1012 = lnj1012 + zrkz1012 + zrky1012 + xysjz1012 + xysjy1012 + dtft1012 + dtrxt1012
sum_1217 = lnj1217 + zrkz1217 + zrky1217 + xysjz1217 + xysjy1217 + dtft1217 + dtrxt1217
sum_1721 = lnj1721 + zrkz1721 + zrky1721 + xysjz1721 + xysjy1721 + dtft1721 + dtrxt1721
sum_2122 = lnj2122 + zrkz2122 + zrky2122 + xysjz2122 + xysjy2122 + dtft2122 + dtrxt2122


def get_week_day(date):
    week_day_dict = {
        1: '星期一',
        2: '星期二',
        3: '星期三',
        4: '星期四',
        5: '星期五',
        6: '星期六',
        0: '星期日',
    }
    day = date.weekday()
    return week_day_dict[day]


wenzi = (yesterday[0:4] + '年' + yesterday[5:7] + '月' +
         yesterday[8:10] + '日' + '(' + get_week_day(datetime.datetime.now()) + ')' +
         '招商花园里总来客' + str(sum_all) + '人次，7-10点' + str(sum_710) + '人次，10-12点' + str(sum_1012) +
         '人次，12-17点' + str(sum_1217) + '人次，17-21点' + str(sum_1721) + '人次，21-22点' + str(sum_2122) +
         '人次。' + '环比（' + huanbi_date[5:7] + '月' + huanbi_date[8:10] + '日客流' + str(int(worksheet.cell_value(40, 7))) + '人次）' + huanbi +
         '%.2f%%' % abs((worksheet.cell_value(41, 9) * 100)) + '。')

sheet.cell(5, 1, wenzi)

wb.save('处理后数据表/花园里客流汇总.xlsx')

print('')
print('写入数据成功======>>>处理后数据表/花园里客流汇总+文字.xlsx')

input()
